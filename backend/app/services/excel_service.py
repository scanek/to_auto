import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_vehicle_excel(vehicle, service_records, fuel_logs, reminders, tyres, documents, analytics) -> bytes:
    wb = openpyxl.Workbook()
    
    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1E293B")
    bold_font = Font(name="Calibri", size=10, bold=True)
    regular_font = Font(name="Calibri", size=10)
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    sub_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    accent_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # -------------------------------------------------------------
    # Sheet 1: Общая сводка (Dashboard)
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Сводка автомобиля"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.merge_cells("A1:D1")
    ws1["A1"] = f"Отчет по автомобилю: {vehicle.make} {vehicle.model} ({vehicle.license_plate or 'Без номера'})"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = Alignment(vertical="center")
    
    purchase_date_str = vehicle.purchase_date.strftime("%d.%m.%Y") if getattr(vehicle, "purchase_date", None) else "—"
    starting_odo_str = f"{vehicle.starting_odometer:,.0f} {vehicle.distance_unit}".replace(",", " ") if getattr(vehicle, "starting_odometer", None) is not None else "0 " + vehicle.distance_unit

    veh_info = [
        ("Марка и модель", f"{vehicle.make} {vehicle.model}"),
        ("Год выпуска", vehicle.year or "—"),
        ("Гос. номер", vehicle.license_plate or "—"),
        ("VIN номер", vehicle.vin or "—"),
        ("Дата покупки / регистрации", purchase_date_str),
        ("Пробег при покупке", starting_odo_str),
        ("Двигатель / КПП", vehicle.engine or "—"),
        ("Спецификация масла", vehicle.oil_spec or "—"),
        ("Текущий пробег", f"{vehicle.current_odometer:,.0f} {vehicle.distance_unit}".replace(",", " ")),
        ("Моточасы", f"{vehicle.current_engine_hours:,.0f} м/ч".replace(",", " ")),
        ("Всего затрат", f"{analytics.total_spend:,.2f} {vehicle.currency}".replace(",", " ")),
        ("Средний расход топлива", f"{analytics.avg_fuel_consumption or 0:.2f} л/100км"),
        ("Стоимость 1 км пути", f"{analytics.cost_per_distance_unit or 0:.2f} {vehicle.currency}/км"),
    ]
    
    row = 3
    for label, val in veh_info:
        ws1.cell(row=row, column=1, value=label).font = bold_font
        ws1.cell(row=row, column=1).fill = sub_fill
        ws1.cell(row=row, column=1).border = thin_border
        
        ws1.cell(row=row, column=2, value=str(val)).font = regular_font
        ws1.cell(row=row, column=2).border = thin_border
        row += 1

    # Categories Breakdown
    row += 2
    ws1.cell(row=row, column=1, value="Структура расходов").font = Font(name="Calibri", size=12, bold=True)
    row += 1
    ws1.cell(row=row, column=1, value="Категория").font = header_font
    ws1.cell(row=row, column=1).fill = header_fill
    ws1.cell(row=row, column=2, value=f"Сумма ({vehicle.currency})").font = header_font
    ws1.cell(row=row, column=2).fill = header_fill
    ws1.cell(row=row, column=3, value="Доля %").font = header_font
    ws1.cell(row=row, column=3).fill = header_fill
    
    for cat in analytics.categories:
        row += 1
        ws1.cell(row=row, column=1, value=cat.category).font = regular_font
        ws1.cell(row=row, column=1).border = thin_border
        ws1.cell(row=row, column=2, value=cat.amount).font = bold_font
        ws1.cell(row=row, column=2).border = thin_border
        ws1.cell(row=row, column=3, value=f"{cat.percentage:.1f}%").font = regular_font
        ws1.cell(row=row, column=3).border = thin_border

    # -------------------------------------------------------------
    # Sheet 2: Журнал ТО, Деталей и Тюнинга
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Журнал ТО и деталей")
    ws2.views.sheetView[0].showGridLines = True
    
    headers_s2 = ["Дата", "Тип", "ТО / Событие", "Пробег (км)", "Моточасы", "Деталь / Позиция", "Бренд", "Артикул", "Кол-во", "Магазин", f"Сумма ({vehicle.currency})"]
    for col_idx, h in enumerate(headers_s2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    r_idx = 2
    for srv in service_records:
        date_str = srv.date.strftime("%d.%m.%Y") if hasattr(srv.date, "strftime") else str(srv.date)[:10]
        type_str = "Плановое ТО" if srv.record_type == "service" else ("Ремонт" if srv.record_type == "repair" else "Тюнинг/Допы")
        
        if srv.items:
            for item in srv.items:
                ws2.cell(row=r_idx, column=1, value=date_str).font = regular_font
                ws2.cell(row=r_idx, column=2, value=type_str).font = regular_font
                ws2.cell(row=r_idx, column=3, value=srv.to_tag or srv.title).font = bold_font
                ws2.cell(row=r_idx, column=4, value=srv.odometer).font = regular_font
                ws2.cell(row=r_idx, column=5, value=srv.engine_hours or "—").font = regular_font
                ws2.cell(row=r_idx, column=6, value=item.name).font = regular_font
                ws2.cell(row=r_idx, column=7, value=item.brand or "—").font = regular_font
                ws2.cell(row=r_idx, column=8, value=item.part_number or "—").font = regular_font
                ws2.cell(row=r_idx, column=9, value=f"{item.quantity} {item.unit}").font = regular_font
                ws2.cell(row=r_idx, column=10, value=item.store or srv.store or "—").font = regular_font
                ws2.cell(row=r_idx, column=11, value=item.total_price).font = bold_font
                for c in range(1, 12):
                    ws2.cell(row=r_idx, column=c).border = thin_border
                r_idx += 1
        else:
            ws2.cell(row=r_idx, column=1, value=date_str).font = regular_font
            ws2.cell(row=r_idx, column=2, value=type_str).font = regular_font
            ws2.cell(row=r_idx, column=3, value=srv.to_tag or srv.title).font = bold_font
            ws2.cell(row=r_idx, column=4, value=srv.odometer).font = regular_font
            ws2.cell(row=r_idx, column=5, value=srv.engine_hours or "—").font = regular_font
            ws2.cell(row=r_idx, column=6, value=srv.title).font = regular_font
            ws2.cell(row=r_idx, column=7, value="—").font = regular_font
            ws2.cell(row=r_idx, column=8, value="—").font = regular_font
            ws2.cell(row=r_idx, column=9, value="1").font = regular_font
            ws2.cell(row=r_idx, column=10, value=srv.store or "—").font = regular_font
            ws2.cell(row=r_idx, column=11, value=srv.total_cost).font = bold_font
            for c in range(1, 12):
                ws2.cell(row=r_idx, column=c).border = thin_border
            r_idx += 1

    # -------------------------------------------------------------
    # Sheet 3: Заправки (Fuel Logs)
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Заправки")
    ws3.views.sheetView[0].showGridLines = True
    
    headers_s3 = ["Дата", "Пробег (км)", f"Литры ({vehicle.fuel_unit})", f"Цена/л ({vehicle.currency})", "Расход (л/100км)", "Полный бак", f"Сумма ({vehicle.currency})", "АЗС"]
    for col_idx, h in enumerate(headers_s3, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    r_idx = 2
    for f in fuel_logs:
        date_str = f.date.strftime("%d.%m.%Y") if hasattr(f.date, "strftime") else str(f.date)[:10]
        ws3.cell(row=r_idx, column=1, value=date_str).font = regular_font
        ws3.cell(row=r_idx, column=2, value=f.odometer).font = regular_font
        ws3.cell(row=r_idx, column=3, value=f.fuel_amount).font = regular_font
        ws3.cell(row=r_idx, column=4, value=f.unit_price).font = regular_font
        ws3.cell(row=r_idx, column=5, value=f.consumption if f.consumption else "—").font = bold_font
        ws3.cell(row=r_idx, column=6, value="Да" if f.is_full_tank else "Нет").font = regular_font
        ws3.cell(row=r_idx, column=7, value=f.total_cost).font = bold_font
        ws3.cell(row=r_idx, column=8, value=f.gas_station or "—").font = regular_font
        for c in range(1, 9):
            ws3.cell(row=r_idx, column=c).border = thin_border
        r_idx += 1

    # -------------------------------------------------------------
    # Sheet 4: Шины и Колеса (Tyres)
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="Шины и Колеса")
    ws4.views.sheetView[0].showGridLines = True
    
    headers_s4 = ["Комплект", "Сезон", "Размерность", "Модель", "Пробег (км)", "Остаток (мм)", "Хранение", "Статус", f"Стоимость ({vehicle.currency})"]
    for col_idx, h in enumerate(headers_s4, 1):
        cell = ws4.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    r_idx = 2
    for t in tyres:
        season_str = "Лето ☀️" if t.season == "summer" else "Зима ❄️"
        status_str = "На автомобиле (Активен)" if t.is_active else "В хранении"
        ws4.cell(row=r_idx, column=1, value=t.name).font = bold_font
        ws4.cell(row=r_idx, column=2, value=season_str).font = regular_font
        ws4.cell(row=r_idx, column=3, value=t.size or "—").font = regular_font
        ws4.cell(row=r_idx, column=4, value=t.brand_model or "—").font = regular_font
        ws4.cell(row=r_idx, column=5, value=t.current_km).font = regular_font
        ws4.cell(row=r_idx, column=6, value=f"{t.tread_depth_mm} мм").font = bold_font
        ws4.cell(row=r_idx, column=7, value=t.storage_location or "—").font = regular_font
        ws4.cell(row=r_idx, column=8, value=status_str).font = bold_font
        ws4.cell(row=r_idx, column=9, value=t.total_price).font = regular_font
        for c in range(1, 10):
            ws4.cell(row=r_idx, column=c).border = thin_border
        r_idx += 1

    # -------------------------------------------------------------
    # Sheet 5: Страховки и Документы (Documents)
    # -------------------------------------------------------------
    ws5 = wb.create_sheet(title="Страховки и Документы")
    ws5.views.sheetView[0].showGridLines = True
    
    headers_s5 = ["Наименование", "Тип", "Компания", "Номер полиса", "Дата начала", "Действует до", f"Стоимость ({vehicle.currency})", "Заметки"]
    for col_idx, h in enumerate(headers_s5, 1):
        cell = ws5.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    r_idx = 2
    for d in documents:
        exp_str = d.expiration_date.strftime("%d.%m.%Y") if d.expiration_date else "Бессрочно"
        start_str = d.issue_date.strftime("%d.%m.%Y") if d.issue_date else "—"
        ws5.cell(row=r_idx, column=1, value=d.title).font = bold_font
        ws5.cell(row=r_idx, column=2, value=d.doc_type).font = regular_font
        ws5.cell(row=r_idx, column=3, value=d.company or "—").font = regular_font
        ws5.cell(row=r_idx, column=4, value=d.document_number or "—").font = regular_font
        ws5.cell(row=r_idx, column=5, value=start_str).font = regular_font
        ws5.cell(row=r_idx, column=6, value=exp_str).font = bold_font
        ws5.cell(row=r_idx, column=7, value=d.price).font = regular_font
        ws5.cell(row=r_idx, column=8, value=d.notes or "—").font = regular_font
        for c in range(1, 9):
            ws5.cell(row=r_idx, column=c).border = thin_border
        r_idx += 1

    # Auto-adjust column widths on all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
